"""One-time historical importer - DB only, no Google Sheets dependency, so the
existing five months of data are safe in SQLite before any credentials exist.

Reuses the per-sheet transaction-list start rows validated in the prior session:
each month sheet's own B18/D18 SUM() formulas define which rows are "official" -
rows above that start are stray/excluded entries the original sheet itself didn't
count, and including them would make our totals disagree with the numbers the user
has been looking at for five months.

Also applies the category cleanup agreed on with the user: ride-hailing and
telecom/utility-recharge transactions that were dumped into the catch-all
"Subscriptions" bucket (or, for electricity, sometimes into "RENT") get split out
into proper Transport / Utilities categories. The original label is preserved in
raw_category for audit - nothing is silently lost, just organized.
"""
import calendar
import re
import sys
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import init_db, session_scope  # noqa: E402
from app.models import SyncStatus, Transaction, TransactionSource, TransactionType  # noqa: E402
from app.repositories.accounts import AccountRepository  # noqa: E402
from app.repositories.categories import CategoryRepository  # noqa: E402
from app.repositories.transactions import TransactionRepository  # noqa: E402
from app.sync.periods import ensure_periods_for_transactions  # noqa: E402
from app.utils import content_hash, next_transaction_id, period_key_for  # noqa: E402

START_ROW = {
    "April 2026":  {"income": 21, "expense": 21},
    "May 2026":    {"income": 19, "expense": 19},
    "June 2026":   {"income": 19, "expense": 20},
    "July 2026":   {"income": 19, "expense": 19},
    "August 2026": {"income": 19, "expense": 19},
}
MONTHS = list(START_ROW.keys())

TRANSPORT_KEYWORDS = ["rapido", "uber"]
UTILITIES_KEYWORDS = ["recharge", "electricity", "broadband"]
# Typo/label-drift consolidation found in the legacy sheet (e.g. a few rows tagged
# "Food" instead of "Food-Order") - same category, just inconsistent labeling.
LABEL_ALIASES = {"food": "Food-Order"}


DATE_IN_DESCRIPTION_RE = re.compile(r"\((\d{1,2})/(\d{1,2})/(\d{2,4})\)")


def extract_date(description: str, expected_year: int, expected_month: int) -> tuple[date, bool]:
    """Most legacy descriptions carry the real transaction date at the end, e.g.
    "Sent to Mom SBI Savings (04/08/26)". Tries DD/MM first (the convention used
    throughout the sheet), falls back to MM/DD, and only accepts a candidate whose
    year/month matches the row's own sheet - a handful of rows have a clearly
    mistyped date (e.g. a "/28/" that fits neither format for that month), and for
    those a fallback is safer than guessing wrong.

    Returns (date, matched) - matched=False tells the caller no real date was found,
    so it can carry forward the previous row's date instead of using the raw fallback
    (paychecks are the exception: they always fall back to the last day of the month,
    matched=False, and are never carry-forward targets or sources)."""
    fallback_day = (
        calendar.monthrange(expected_year, expected_month)[1]
        if description.strip().lower().startswith("paycheck") else 1
    )
    fallback = date(expected_year, expected_month, fallback_day)
    m = DATE_IN_DESCRIPTION_RE.search(description.strip())
    if not m:
        return fallback, False
    a, b, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    year = yy + 2000 if yy < 100 else yy
    candidates = []
    if 1 <= a <= 31 and 1 <= b <= 12:
        candidates.append((b, a))  # DD/MM
    if 1 <= b <= 31 and 1 <= a <= 12 and a != b:
        candidates.append((a, b))  # MM/DD
    for month, day in candidates:
        if year == expected_year and month == expected_month:
            try:
                return date(year, month, day), True
            except ValueError:
                continue
    return fallback, False


def reclassify(description: str, original_category: str) -> str:
    base = LABEL_ALIASES.get((original_category or "").strip().lower(), original_category or "Other")
    d = description.lower()
    if any(k in d for k in TRANSPORT_KEYWORDS):
        return "Transport"
    if any(k in d for k in UTILITIES_KEYWORDS):
        return "Utilities"
    return base


MONTH_NUM = {"April": 4, "May": 5, "June": 6, "July": 7, "August": 8}


def _resolve_with_carry_forward(description: str, year: int, month_num: int, last_date: Optional[date]) -> date:
    """Wraps extract_date(): an unmatched, non-paycheck row inherits the previous
    row's date in the same (month, type) sequence - e.g. "Cold Drink" with no date
    of its own, immediately after "Zomato (07/05/26)", is almost certainly the same
    day. Falls back to extract_date()'s own fallback if there's no prior date yet."""
    d, matched = extract_date(description, year, month_num)
    if not matched and last_date is not None and not description.strip().lower().startswith("paycheck"):
        return last_date
    return d


def extract(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = []
    for month in MONTHS:
        ws = wb[month]
        start = START_ROW[month]
        maxr = ws.max_row
        month_name, year_str = month.split()
        year, month_num = int(year_str), MONTH_NUM[month_name]

        r = start["income"]
        last_date = None
        while r <= maxr:
            desc, amt = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
            if desc is not None and amt is not None:
                d = _resolve_with_carry_forward(str(desc), year, month_num, last_date)
                last_date = d
                rows.append(dict(month=month, type="Income", raw_category="Income",
                                  category="Income", description=str(desc), amount=float(amt), date=d))
            r += 1

        r = start["expense"]
        last_date = None
        while r <= maxr:
            desc, amt, cat = (ws.cell(row=r, column=3).value, ws.cell(row=r, column=4).value,
                               ws.cell(row=r, column=5).value)
            if desc is not None and amt is not None:
                raw_cat = (cat or "Other").strip()
                final_cat = reclassify(str(desc), raw_cat)
                d = _resolve_with_carry_forward(str(desc), year, month_num, last_date)
                last_date = d
                rows.append(dict(month=month, type="Expense", raw_category=raw_cat,
                                  category=final_cat, description=str(desc), amount=float(amt), date=d))
            r += 1
    return rows


def run_seed(xlsx_path: str) -> dict:
    init_db()
    rows = extract(xlsx_path)

    with session_scope() as session:
        cat_repo = CategoryRepository(session)
        acct_repo = AccountRepository(session)
        tx_repo = TransactionRepository(session)
        cat_repo.ensure_defaults()
        account = acct_repo.ensure_default()

        existing_ids = tx_repo.all_transaction_ids()
        created = 0
        reclassified = 0

        for row in rows:
            # date was already resolved in extract() (real day when the description
            # has one and it matches the row's own month, carried forward from the
            # previous row when it doesn't, last-day-of-month for paychecks, else the
            # 1st) - period_key below still pins to the sheet's own month regardless,
            # so the validated monthly totals are never affected by any of this.
            d = row["date"]

            category = cat_repo.add(row["category"])
            if row["category"] != row["raw_category"]:
                reclassified += 1

            tid = next_transaction_id(existing_ids, d.year)
            existing_ids.append(tid)

            h = content_hash(date=d, description=row["description"], amount=row["amount"],
                              transaction_type=row["type"], category_name=category.name,
                              account_name=account.name, notes=None)
            txn = Transaction(
                transaction_id=tid, date=d, description=row["description"].strip(),
                amount=round(row["amount"], 2), transaction_type=TransactionType(row["type"]),
                category_id=category.id, account_id=account.id, period_key=period_key_for(d),
                raw_category=row["raw_category"], source=TransactionSource.legacy_import,
                content_hash=h, sync_status=SyncStatus.pending,
            )
            session.add(txn)
            created += 1

        session.flush()

        ensure_periods_for_transactions(session)

    return {"created": created, "reclassified": reclassified, "total_rows": len(rows)}


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "../../Monthly Budget Sagnik Das.xlsx"
    result = run_seed(path)
    print(result)
